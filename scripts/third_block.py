import pandas as pd
from openpyxl import load_workbook
import os

# Ask the user for the input Excel file and MG CSV file
excel_file = input("Enter the path of the input Excel file (e.g., AMPERSAND_OCTOBER_MIS_Summary_2024.xlsx): ")

# Ensure the excel_file path is absolute
excel_file = os.path.abspath(excel_file)

# Check if the file exists, if not, try to find a similar file
if not os.path.exists(excel_file):
    # Get the directory of the script or current working directory
    script_dir = os.path.dirname(excel_file) or os.getcwd()
    
    # List all Excel files in the directory
    excel_files = [f for f in os.listdir(script_dir) if f.endswith('.xlsx')]
    
    if excel_files:
        # Choose the first Excel file if multiple exist
        excel_file = os.path.join(script_dir, excel_files[0])
        print(f"Using file: {excel_file}")
    else:
        print("No Excel file found!")
        exit()

mg_file = input("Enter the path of the MG CSV file (e.g., MG.csv): ")

# Load the MG CSV file
mg_csv = pd.read_csv(mg_file)

# Load the Excel workbook
workbook = load_workbook(excel_file)

# Get the cross_dock_name and MG column from the CSV
mg_mapping = dict(zip(mg_csv['cross_dock_name'], mg_csv['MG']))

# Iterate through the sheets in the Excel workbook
for sheet_name in workbook.sheetnames:
    if sheet_name in mg_mapping:
        # Access the sheet
        sheet = workbook[sheet_name]
        
        # Update cell K4 with the corresponding MG value
        sheet['K4'] = mg_mapping[sheet_name]
        
        # Set the formula in K4 to multiply J4 by the MG value
        sheet['K4'] = f'=J4*{mg_mapping[sheet_name]}'

# Save the updated workbook
output_file = excel_file.replace(".xlsx", "_Updated.xlsx")
workbook.save(output_file)
print(f"MG values updated successfully! The updated file is saved as {output_file}")