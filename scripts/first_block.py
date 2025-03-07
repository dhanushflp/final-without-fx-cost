import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import calendar

# Prompt for file path
file_path = input("Enter the CSV file path (e.g., ZEVO_OCT.csv): ")

try:
    # Load data
    data = pd.read_csv(file_path, low_memory=False)
except FileNotFoundError:
    print("Error: File not found. Please provide a valid file path.")
    exit()

# Prompt for the vendor and month
vendor_name = input("Enter vendor name: ").upper()
month_name = input("Enter month: ").upper()
year_name = input("Enter year: ")

# Generate the output file name based on vendor and month
output_file = f"{vendor_name}_{month_name}_MIS_Summary_{year_name}.xlsx"

# Automatically extract unique cross-dock names from column 'U'
cross_dock_names = data['cross_dock_name'].unique().tolist()

# Display cross-dock names line by line
print("Available cross-docks:")
for name in cross_dock_names:
    print(name)
print("-------------------------------------")

# You can now remove the prompt for the number of cross-docks and just work with all available cross-docks
# Filter data for all cross-dock names
filtered_data = data[data['cross_dock_name'].isin(cross_dock_names)]

# Create an Excel file with the header and calculations
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for name in cross_dock_names:
        # Filter for each cross-dock name
        subset = filtered_data[filtered_data['cross_dock_name'] == name]

        # Create pivot table
        pivot_table = subset.pivot_table(
            index='sheet_create_date',
            columns='sheet_shipment_last_status',
            values='order_id',
            aggfunc='count'
        )

        # Convert the index to datetime format, keep only the date, and sort
        pivot_table.index = pd.to_datetime(pivot_table.index, errors='coerce').date
        pivot_table = pivot_table.sort_index()

        # Add Grand Total row
        pivot_table.loc['Grand Total'] = pivot_table.sum()

        # Write the pivot table to a new sheet named after cross_dock_name
        pivot_table.to_excel(writer, sheet_name=name)

        # Add custom headers and calculations (details omitted for brevity, same as original logic)
        workbook = writer.book
        worksheet = workbook[name]

        # Add headers in the specific cells (as per your request)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Set header values and background color
        headers = {
            "I3": "DELIVERED+PARTIALLY_DELIVERED",
            "J3": "Landing Plan",
            "K3": "Mg on LP",
            "L3": "Billable order",
            "M3": "Slab",
            "N3": "Rate",
            "O3": "Amount",
            "Q3": "cross_dock_name",
            "R3": "DELIVERED",
            "S3": "PARTIALLY_DELIVERED",
            "T3": "RTO",
            "U3": "UNDELIVERED",
            "V3": "Post-Slot Breach",
            "W3": "Slab",
            "X3": "MG Invoice Amt",
            "Y3": "DNE Amount",
            "Z3": "Final Invoice amount",
            "P9": "OXD City",
            "Q9": "DELIVERED",
            "R9": "PRTO",
            "S9": "Total RTO (2.5%)",
            "T9": "RA(3%)",
            "U9": "Post Breach (5%)",
            "V9": "Remarks",
            "W9": "RTO Penalty",
            "X9": "RA Penalty",
            "Y9": "PB Penalty",
            "Z9": "Final Penalty",
            "P16": "DELIVERED",
            "Q16": "PARTIALLY_DELIVERED",
            "R16": "Slab",
            "S16": "Rate",
            "T16": "Base Invoice value"
        }

        for cell, value in headers.items():
            worksheet[cell] = value
            worksheet[cell].fill = yellow_fill
            worksheet[cell].font = bold_font
            worksheet[cell].alignment = center_alignment

        # Get the month and year from the sheet_create_date (for the first row in the subset)
        sheet_create_date = pd.to_datetime(subset['sheet_create_date'].iloc[0])
        month_days = calendar.monthrange(sheet_create_date.year, sheet_create_date.month)[1]  # Get number of days in the month

        # Count the occurrences of different statuses in the 'sheet_shipment_last_status' column
        status_counts = subset['sheet_shipment_last_status'].value_counts()

        # Extract counts for DELIVERED, PARTIALLY_DELIVERED, RTO, SYSTEM_RTO, and OUT_FOR_DELIVERY
        delivered_count = status_counts.get('DELIVERED', 0)
        out_for_delivery_count = status_counts.get('OUT_FOR_DELIVERY', 0)
        partially_delivered_count = status_counts.get('PARTIALLY_DELIVERED', 0)
        undelivered_count = status_counts.get('UNDELIVERED', 0)
        rto_count = status_counts.get('RTO', 0)
        system_rto_count = status_counts.get('SYSTEM_RTO', 0)

        # Count occurrences of "Post-Slot Breach" in 'slot_breach_flag' column
        post_slot_breach_count = subset[subset['slot_breach_flag'] == 'Post-Slot Breach'].shape[0]

        # Add calculations in the cells based on the pivot data
        worksheet["I4"] = "=SUM(R4:S4)"  # Blank cell
        worksheet["J4"] = ""  # Blank cell
        worksheet["K4"] = ""  # Blank cell
        worksheet["L4"] = "=MAX(I4,K4)"

        # Use the dynamic number of days in the month for the formula in M4
        worksheet["M4"] = f"=L4/{month_days}"

        worksheet["N4"] = ""  # Blank cell
        worksheet["O4"] = "=L4*N4"

        # Cross-dock specific values
        worksheet["Q4"] = f"{name}"  # Cross-dock name

        # Set the correct values for DELIVERED, PARTIALLY_DELIVERED, RTO, SYSTEM_RTO in the worksheet
        worksheet["R4"] = delivered_count + out_for_delivery_count  # Sum DELIVERED and OUT_FOR_DELIVERY counts
        worksheet["S4"] = partially_delivered_count  # Count of PARTIALLY_DELIVERED
        worksheet["T4"] = rto_count + system_rto_count  # Count of RTO + SYSTEM_RTO
        worksheet["U4"] = undelivered_count

        worksheet["V4"] = post_slot_breach_count  # Set the count of Post-Slot Breach in V4

        worksheet["W4"] = "=M4"
        worksheet["X4"] = "=O4"
        worksheet["Y4"] = ""  # Blank cell
        worksheet["Z4"] = "=X4-Z10"

        # Cells for penalty calculations
        worksheet["P10"] = "=Q4"
        worksheet["Q10"] = "=R4"
        worksheet["R10"] = "=S4"
        worksheet["S10"] = "=ROUND(T4/SUM(R4:U4),2)"
        worksheet["T10"] = "=ROUND(U4/SUM(R4:U4), 2)"
        worksheet["U10"] = "=ROUND(V4/SUM(R4:U4), 2)"
        worksheet["V10"] = ""  # Blank cell
        worksheet["W10"] = (
               '=IF(AND(S10>2.5%, S10<=3.5%), T17*2%, '
               'IF(AND(S10>3.5%, S10<=4.5%), T17*3%, '
               'IF(S10>4.5%, T17*5%, 0)))'
        )

        worksheet["X10"] = (
            '=IF(AND(T10>3%, T10<=4%), T17*1%, '
            'IF(AND(T10>4%, T10<=5%), T17*2%, '
            'IF(T10>5%, T17*4%, 0)))'
        )

        worksheet["Y10"] = (
            '=IF(AND(U10>5%, U10<=6%), S17*V4*3%, '
            'IF(AND(U10>6%, U10<=8%), S17*V4*5%, '
            'IF(U10>8%, S17*V4*8%, 0)))'
        )

        worksheet["Z10"] = "=SUM(W10:Y10)"

        # Cells for final calculation
        worksheet["P17"] = "=R4"
        worksheet["Q17"] = "=S4"
        worksheet["R17"] = f"=SUM(P17:Q17)/{month_days}"  # Adjust number of days dynamically
        worksheet["S17"] = ""  # Blank cell
        worksheet["T17"] = "=SUM(P17:Q17)*S17"

        # Auto-adjust column width and center-align the text
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save the workbook after processing all cross-docks
    workbook.save(output_file)

    print(f"Data for {vendor_name} has been processed and saved to {output_file}.")
