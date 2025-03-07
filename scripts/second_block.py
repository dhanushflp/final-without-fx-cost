import pandas as pd
from openpyxl import load_workbook

# User input for file paths
input_excel = input("Enter the path to the input Excel file (e.g., AMPERSAND_OCTOBER_MIS_Summary_2024.xlsx): ")
landing_csv = input("Enter the path to the landing plan CSV file (e.g., landingplan.csv): ")

# Load the landing plan CSV
landing_data = pd.read_csv(landing_csv)

# Clean the LP column: Remove commas, fill NaN with 0, and convert to integers
landing_data["LP"] = (
    landing_data["LP"]
    .replace({",": ""}, regex=True)  # Remove commas
    .fillna(0)  # Replace NaN with 0
    .astype(int)  # Convert to integers
)

# Load the Excel file
wb = load_workbook(input_excel)

# Iterate through each sheet in the Excel file
for sheet_name in wb.sheetnames:
    # Check if the sheet name exists in the landing plan data
    if sheet_name in landing_data["cross_dock_name"].values:
        # Filter landing plan data for the current sheet name
        sheet_data = landing_data[landing_data["cross_dock_name"] == sheet_name]
        # Calculate the sum of the LP column
        total_lp = sheet_data["LP"].sum()
        # Open the corresponding sheet in the Excel file
        sheet = wb[sheet_name]
        # Write the sum to cell J4
        sheet["J4"] = total_lp

# Save the updated Excel file
wb.save(input_excel)

print(f"Sum of LP values updated successfully in J4 for each sheet in {input_excel}.")
